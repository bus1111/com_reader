import os
import sys
import time
import serial
from serial.tools import list_ports
from datetime import datetime
from threading import Thread
from openpyxl import Workbook


PREFIX = b'sendordata'


def create_workbook(sensor_count):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сенсоры'
    ws.cell(1, 1, 'Время')
    for i in range(1, sensor_count + 1):
        ws.cell(row=1, column=i + 1, value=f'Сенсор {i}')
    return wb


def reader(wb, port, sensor_count, filename, should_reset, should_stop):
    ws = wb.active
    last_save = 0
    with serial.Serial(baudrate=115200, timeout=0.2) as ser:
        ser.port = port.device
        if sys.platform != 'win32':
            # Особенности Windows драйвера
            should_reset = not should_reset
        if not should_reset:
            ser.rts = False
            ser.dtr = False
        ser.open()
        while True:
            try:
                line = ser.readline()
                if should_stop() or not ser.is_open:
                    break
                if len(line) == 0 or not line.startswith(PREFIX):
                    continue
            except Exception as e:
                print('Ошибка чтения, нажмите Enter:', e)
                break

            try:
                line = line.strip().replace(PREFIX, b'').decode('utf-8')
                values = line.split('/')
                print(f'Получено: {", ".join(values):{sensor_count*8}}')
                sys.stdout.flush()
                values = [int(v) for v in values[:sensor_count]]
                ws.append([datetime.now()] + values)
                if time.time() - last_save > 10:
                    wb.save(filename)
                    last_save = time.time()
            except Exception as e:
                print('Ошибка обработки:', e)


def menu():
    while True:
        ports = sorted(list_ports.comports())
        if len(ports) == 0:
            print('Устройства не найдены, повторите.')
            input()
            continue
        break

    print('Доступные устройства:')
    for i, port in enumerate(ports):
        print(f'   * {i + 1}: {port.name} - {port.description}')

    while True:
        try:
            selected_port = 1
            selected_port = input(f'Выберите устройство [{selected_port}]: ') or selected_port
            selected_port = int(selected_port)
            if selected_port > len(ports) or selected_port <= 0:
                print('Неверный ввод')
            else:
                break
        except Exception:
            print('Неверный ввод')
            continue

    while True:
        try:
            sensor_count = 3
            sensor_count = input(f'Введите количество сенсоров [{sensor_count}]: ') or sensor_count
            sensor_count = int(sensor_count)
            break
        except Exception:
            print('Неверный ввод')
            continue
    should_reset = input(f'Cбросить устройство перед началом (да/нет) [нет]: ').lower()
    should_reset = should_reset.startswith('д') or should_reset.startswith('y')

    return ports[selected_port - 1], sensor_count, should_reset


if __name__ == '__main__':
    while True:
        port, sensor_count, should_reset = menu()
        wb = create_workbook(sensor_count)
        stop_signal = False
        filename = f'{datetime.strftime(datetime.now(), "Сенсоры %d %b %Y %H-%M-%S.xlsx")}'
        t = Thread(target=reader, args=(wb, port, sensor_count, filename, should_reset, lambda: stop_signal))
        t.start()
        input('Запись начата, нажмите Enter чтобы остановить\n')
        stop_signal = True
        print('Запись останавливается...')
        t.join(timeout=5)
        wb.save(filename)
        print(f'Запись сохранена в файл {os.path.abspath(filename)}')
        input('Нажмите Enter чтобы начать новую запись')
